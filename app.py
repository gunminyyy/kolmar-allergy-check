import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageChops
import io
import re
import os
import fitz  # PyMuPDF
import numpy as np
import gc
import math
from datetime import datetime
import openpyxl.utils

# 1. 페이지 설정
st.set_page_config(page_title="MSDS 스마트 변환기", layout="wide")
st.title("MSDS 양식 변환기")
st.markdown("---")

# --------------------------------------------------------------------------
# [1. 유틸리티 함수]
# --------------------------------------------------------------------------
FONT_STYLE = Font(name='굴림', size=8)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

def safe_write_force(ws, row, col, value, center=False):
    cell = ws.cell(row=row, column=col)
    try: cell.value = value
    except AttributeError:
        try:
            for rng in list(ws.merged_cells.ranges):
                if cell.coordinate in rng:
                    ws.unmerge_cells(str(rng))
                    cell = ws.cell(row=row, column=col)
                    break
            cell.value = value
        except: pass
    if cell.font.name != '굴림': cell.font = FONT_STYLE
    if center: cell.alignment = ALIGN_CENTER
    else: cell.alignment = ALIGN_LEFT

def get_description_smart(code, code_map):
    clean_code = str(code).replace(" ", "").upper().strip()
    if clean_code in code_map: return code_map[clean_code]
    if "+" in clean_code:
        parts = clean_code.split("+")
        found_texts = []
        for p in parts:
            if p in code_map: found_texts.append(code_map[p])
        if found_texts: return " ".join(found_texts)
    return ""

def calculate_smart_height_basic(text, mode="CFF(K)"): 
    if not text: return 19.2
    
    lines = str(text).split('\n')
    total_visual_lines = 0
    
    if "E" in mode:
        char_limit = 65.0
        for line in lines:
            if len(line) == 0:
                total_visual_lines += 1
            else:
                words = line.split(" ")
                current_len = 0
                lines_for_this_paragraph = 1
                
                for word in words:
                    if current_len == 0:
                        current_len = len(word)
                    elif current_len + 1 + len(word) <= char_limit:
                        current_len += 1 + len(word)
                    else:
                        lines_for_this_paragraph += 1
                        current_len = len(word)
                total_visual_lines += lines_for_this_paragraph
                
        if total_visual_lines <= 1: 
            return 18.75
        elif total_visual_lines == 2: 
            return 25.5
        elif total_visual_lines == 3: 
            return 36.0
        elif total_visual_lines == 4: 
            return 44.0
        elif total_visual_lines == 5: 
            return 54.0
        else: 
            return 64.0 + (total_visual_lines - 6) * 10.0
    else:
        char_limit = 45.0
        for line in lines:
            if len(line) == 0:
                total_visual_lines += 1
            else:
                total_visual_lines += math.ceil(len(line) / char_limit)
                
        if total_visual_lines <= 1: 
            return 19.2
        elif total_visual_lines == 2: 
            return 26.0
        elif total_visual_lines == 3: 
            return 36.0
        else: 
            return 45.0

def format_and_calc_height_sec47(text, mode="CFF(K)"):
    if not text: return "", 19.2
    
    if "E" in mode:
        keywords = r"(IF|If|Get|When|Wash|Remove|Take|Prevent|Call|Move|Settle|Please|After|Should|Rescuer|For|Do|Wipe|Follow|Stop|Collect|Make|Absorb|Put|Since|Contaminated|Without|Empty|Keep|Store|The|It|Some|During|Containers)"
        formatted_text = re.sub(r'(?<=[a-z0-9\)\]\.\;])\s+(' + keywords + r'\b)', r'\n\1', text)
        formatted_text = re.sub(r'\.([A-Z])', r'.\n\1', formatted_text)
        
        formatted_text = formatted_text.replace("Follow\nStop", "Follow Stop")
        
        lines = [line.strip() for line in formatted_text.split('\n') if line.strip()]
        final_text = "\n".join(lines)
        
        char_limit_per_line = 73.0
        total_visual_lines = 0
        for line in lines:
            if len(line) == 0:
                total_visual_lines += 1
            else:
                words = line.split(" ")
                current_len = 0
                lines_for_this_paragraph = 1
                
                for word in words:
                    if current_len == 0:
                        current_len = len(word)
                    elif current_len + 1 + len(word) <= char_limit_per_line:
                        current_len += 1 + len(word)
                    else:
                        lines_for_this_paragraph += 1
                        current_len = len(word)
                total_visual_lines += lines_for_this_paragraph
            
        if total_visual_lines == 0: total_visual_lines = 1
        height = total_visual_lines * 12.0
        
        if height < 24.0: height = 24.0
        
        return final_text, height
    else:
        formatted_text = re.sub(r'(?<!\d)\.(?!\d)(?!\n)', '.\n', text)
        lines = [line.strip() for line in formatted_text.split('\n') if line.strip()]
        final_text = "\n".join(lines)
        
        char_limit_per_line = 45
        total_visual_lines = 0
        for line in lines:
            line_len = 0
            for ch in line:
                line_len += 2 if '가' <= ch <= '힣' else 1.1 
            visual_lines = math.ceil(line_len / (char_limit_per_line * 2)) 
            if visual_lines == 0: visual_lines = 1
            total_visual_lines += visual_lines
        if total_visual_lines == 0: total_visual_lines = 1
        height = (total_visual_lines * 10) + 10
        if height < 24.0: height = 24.0
        return final_text, height

def fill_fixed_range(ws, start_row, end_row, codes, code_map, mode="CFF(K)"):
    unique_codes = []; seen = set()
    for c in codes:
        clean = c.replace(" ", "").upper().strip()
        if clean not in seen: unique_codes.append(clean); seen.add(clean)
    limit = end_row - start_row + 1
    for i in range(limit):
        current_row = start_row + i
        if i < len(unique_codes):
            code = unique_codes[i]
            desc = get_description_smart(code, code_map)
            ws.row_dimensions[current_row].hidden = False
            final_height = calculate_smart_height_basic(desc, mode)
            ws.row_dimensions[current_row].height = final_height
            safe_write_force(ws, current_row, 2, code, center=False)
            safe_write_force(ws, current_row, 4, desc, center=False)
        else:
            if "K" in mode and current_row in [25, 38, 50, 64, 70]:
                ws.row_dimensions[current_row].hidden = False
                safe_write_force(ws, current_row, 2, "")
                safe_write_force(ws, current_row, 4, "자료없음", center=False)
            elif "E" in mode and current_row in [24, 38, 50, 64, 70]:
                ws.row_dimensions[current_row].hidden = False
                safe_write_force(ws, current_row, 2, "")
                safe_write_force(ws, current_row, 4, "no data available", center=False)
            else:
                ws.row_dimensions[current_row].hidden = True
                safe_write_force(ws, current_row, 2, "") 
                safe_write_force(ws, current_row, 4, "")

def fill_composition_data(ws, comp_data, cas_to_name_map, mode="CFF(K)"):
    start_row = 80; end_row = 123
    if "E" in mode: end_row = 122
    limit = end_row - start_row + 1

    for i in range(limit):
        current_row = start_row + i
        if i < len(comp_data):
            cas_no, concentration = comp_data[i]
            clean_cas = cas_no.replace(" ", "").strip()
            chem_name = cas_to_name_map.get(clean_cas, "")
            
            ws.row_dimensions[current_row].hidden = False
            ws.row_dimensions[current_row].height = 26.7
            
            safe_write_force(ws, current_row, 1, chem_name, center=False)
            safe_write_force(ws, current_row, 4, cas_no, center=False) 
            safe_write_force(ws, current_row, 6, concentration if concentration else "", center=True)
        else:
            ws.row_dimensions[current_row].hidden = True
            safe_write_force(ws, current_row, 1, "")
            safe_write_force(ws, current_row, 4, "")
            safe_write_force(ws, current_row, 6, "")

def fill_regulatory_section(ws, start_row, end_row, substances, data_map, col_key, mode="CFF(K)"):
    limit = end_row - start_row + 1
    for i in range(limit):
        current_row = start_row + i
        if i < len(substances):
            substance_name = substances[i]
            safe_write_force(ws, current_row, 1, substance_name, center=False)
            cell_data = ""
            if substance_name in data_map:
                cell_data = str(data_map[substance_name].get(col_key, ""))
                if cell_data == "nan": cell_data = ""
            safe_write_force(ws, current_row, 2, cell_data, center=False)
            ws.row_dimensions[current_row].hidden = False
            _, h = format_and_calc_height_sec47(cell_data, mode=mode)
            if h < 24.0: h = 24.0 
            ws.row_dimensions[current_row].height = h
        else:
            safe_write_force(ws, current_row, 1, "")
            safe_write_force(ws, current_row, 2, "")
            ws.row_dimensions[current_row].hidden = True

# --------------------------------------------------------------------------
# [2. 이미지 함수]
# --------------------------------------------------------------------------
def auto_crop(pil_img):
    try:
        if pil_img.mode != 'RGB':
            bg = PILImage.new('RGB', pil_img.size, (255, 255, 255))
            if pil_img.mode == 'RGBA': bg.paste(pil_img, mask=pil_img.split()[3])
            else: bg.paste(pil_img)
            pil_img = bg
        bbox = ImageChops.invert(pil_img).getbbox()
        if bbox: return pil_img.crop(bbox)
        return pil_img
    except: return pil_img

def normalize_image_legacy(pil_img):
    try:
        if pil_img.mode in ('RGBA', 'LA') or (pil_img.mode == 'P' and 'transparency' in pil_img.info):
            background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
            if pil_img.mode == 'P': pil_img = pil_img.convert('RGBA')
            background.paste(pil_img, mask=pil_img.split()[3])
            pil_img = background
        else: pil_img = pil_img.convert('RGB')
        return pil_img.resize((32, 32)).convert('L')
    except: return pil_img.resize((32, 32)).convert('L')

def normalize_image_smart(pil_img):
    try:
        cropped_img = auto_crop(pil_img)
        return cropped_img.resize((64, 64)).convert('L')
    except: return pil_img.resize((64, 64)).convert('L')

def get_reference_images():
    img_folder = "reference_imgs"
    if not os.path.exists(img_folder): return {}, False
    try:
        ref_images = {}
        file_list = sorted(os.listdir(img_folder)) 
        for fname in file_list:
            if fname.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.tif', '.tiff')):
                full_path = os.path.join(img_folder, fname)
                try:
                    pil_img = PILImage.open(full_path)
                    ref_images[fname] = pil_img
                except: continue
        return ref_images, True
    except: return {}, False

def is_blue_dominant(pil_img):
    try:
        img = pil_img.resize((50, 50)).convert('RGB')
        data = np.array(img)
        r = data[:,:,0].astype(int); g = data[:,:,1].astype(int); b = data[:,:,2].astype(int)
        blue_mask = (b > r + 30) & (b > g + 30)
        blue_ratio = np.sum(blue_mask) / (50 * 50)
        return blue_ratio > 0.05
    except: return False

def is_square_shaped(width, height):
    if height == 0: return False
    ratio = width / height
    return 0.8 < ratio < 1.2 

def find_best_match_name(src_img, ref_images, mode="CFF(K)"):
    best_score = float('inf')
    best_name = None
    
    if mode in ["HP(K)", "HP(E)"]:
        src_norm = normalize_image_smart(src_img)
        threshold = 70 
    else: 
        src_norm = normalize_image_legacy(src_img)
        threshold = 65

    try:
        src_arr = np.array(src_norm, dtype='int16')
        for name, ref_img in ref_images.items():
            if mode in ["HP(K)", "HP(E)"]: 
                ref_norm = normalize_image_smart(ref_img)
            else: 
                ref_norm = normalize_image_legacy(ref_img)
                
            ref_arr = np.array(ref_norm, dtype='int16')
            diff = np.mean(np.abs(src_arr - ref_arr))
            if diff < best_score:
                best_score = diff
                best_name = name
        
        if best_score < threshold: return best_name
        else: return None
    except: return None

def extract_number(filename):
    nums = re.findall(r'\d+', filename)
    return int(nums[0]) if nums else 999

def extract_codes_ordered(text):
    regex_code = re.compile(r"([HP]\s?\d{3}(?:\s*\+\s*[HP]\s?\d{3})*)")
    matches = regex_code.findall(text)
    res = []
    seen = set()
    for c_raw in matches:
        c = c_raw.replace(" ", "").upper()
        if c not in seen:
            seen.add(c)
            res.append(c)
    return res

# --------------------------------------------------------------------------
# [3. 파서 함수]
# --------------------------------------------------------------------------
def get_clustered_lines(doc):
    all_lines = []
    noise_regexs = [
        r'^\s*\d+\s*/\s*\d+\s*$', r'물질안전보건자료', r'Material Safety Data Sheet', 
        r'PAGE', r'Ver\.\s*:?\s*\d+\.?\d*', r'발행일\s*:?.*', r'Date of issue',
        r'주식회사\s*고려.*', r'Cff', r'Corea\s*flavors.*', r'제\s*품\s*명\s*:?.*',
        r'according to the Global Harmonized System', r'Product Name', 
        r'Date\s*:\s*\d{2}\.[a-zA-Z]{3}\.\d{4}'
    ]
    global_y_offset = 0
    for page in doc:
        page_h = page.rect.height
        clip_rect = fitz.Rect(0, 60, page.rect.width, page_h - 50)
        words = page.get_text("words", clip=clip_rect)
        words.sort(key=lambda w: w[1]) 
        rows = []
        if words:
            current_row = [words[0]]
            row_base_y = words[0][1]
            for w in words[1:]:
                if abs(w[1] - row_base_y) < 8:
                    current_row.append(w)
                else:
                    current_row.sort(key=lambda x: x[0])
                    rows.append(current_row)
                    current_row = [w]
                    row_base_y = w[1]
            if current_row:
                current_row.sort(key=lambda x: x[0])
                rows.append(current_row)
        for row in rows:
            line_text = " ".join([w[4] for w in row])
            is_noise = False
            for pat in noise_regexs:
                if re.search(pat, line_text, re.IGNORECASE):
                    is_noise = True; break
            if not is_noise:
                avg_y = sum([w[1] for w in row]) / len(row)
                all_lines.append({
                    'text': line_text,
                    'global_y0': avg_y + global_y_offset,
                    'global_y1': (sum([w[3] for w in row]) / len(row)) + global_y_offset
                })
        global_y_offset += page_h
    return all_lines

def extract_section_smart(all_lines, start_kw, end_kw, mode="CFF(K)"):
    start_idx = -1; end_idx = -1
    clean_start_kw = start_kw.replace(" ", "")
    for i, line in enumerate(all_lines):
        if "E" in mode:
            if clean_start_kw.lower() in line['text'].replace(" ", "").lower():
                start_idx = i; break
        else:
            if clean_start_kw in line['text'].replace(" ", ""):
                start_idx = i; break
    if start_idx == -1: return ""
    
    if isinstance(end_kw, str): end_kw = [end_kw]
    clean_end_kws = [k.replace(" ", "") for k in end_kw]
    
    for i in range(start_idx + 1, len(all_lines)):
        line_clean = all_lines[i]['text'].replace(" ", "")
        for cek in clean_end_kws:
            if "E" in mode:
                if cek.lower() in line_clean.lower(): end_idx = i; break
            else:
                if cek in line_clean: end_idx = i; break
        if end_idx != -1: break
    if end_idx == -1: end_idx = len(all_lines)
    
    target_lines_raw = all_lines[start_idx : end_idx]
    if not target_lines_raw: return ""
    
    first_line = target_lines_raw[0].copy()
    txt = first_line['text']
    escaped_kw = re.escape(start_kw)
    pattern_str = escaped_kw.replace(r"\ ", r"\s*")
    match = re.search(pattern_str, txt, re.IGNORECASE)
    if match:
        content_part = txt[match.end():].strip()
        content_part = re.sub(r"^[:\.\-\s]+", "", content_part)
        first_line['text'] = content_part
    else:
        if start_kw in txt:
            parts = txt.split(start_kw, 1)
            first_line['text'] = parts[1].strip() if len(parts) > 1 else ""
        else: first_line['text'] = ""
    
    target_lines = []; 
    if first_line['text'].strip(): target_lines.append(first_line)
    target_lines.extend(target_lines_raw[1:])
    if not target_lines: return ""
    
    if mode == "CFF(E)":
        garbage_heads = [
            "Classification of the substance or mixture", "Classification of the substance or", "mixture",
            "Precautionary statements", "Hazard pictograms", "Signal word", 
            "Hazard statements", "Response", "Storage", "Disposal", "Other hazards",
            "General advice", "In case of eye contact", "In case of skin contact", "If inhaled", "If swallowed",
            "Special note for doctors", "Extinguishing media", "Special hazards arising from the", "Advice for firefighters",
            "Personal precautions, protective", "Environmental precautions", "Methods and materials for containment",
            "Precautions for safe handling", "Conditions for safe storage, including",
            "Internal regulations", "ACGIH regulations", "Biological exposure standards",
            "arising from the", ", protective", "precautions", "and materials for containment", 
            "for safe handling", "for safe storage, including", "conditions for safe storage, including"
        ]
        sensitive_garbage_regex = []
    elif mode == "HP(E)":
        garbage_heads = [
            "Classification of the substance or mixture", "Classification of the substance or", "mixture",
            "Precautionary statements", "Hazard pictograms", "Signal word", 
            "Hazard statements", "Response", "Storage", "Disposal", "Other hazards",
            "General advice", "In case of eye contact", "In case of skin contact", "If inhaled", "If swallowed",
            "Special note for doctors", "Extinguishing media", "Special hazards arising from the", "Advice for firefighters",
            "Personal precautions, protective", "Environmental precautions", "Methods and materials for containment",
            "Precautions for safe handling", "Conditions for safe storage, including",
            "Internal regulations", "ACGIH regulations", "Biological exposure standards",
            "arising from the", ", protective", "precautions", "and materials for containment", 
            "for safe handling", "for safe storage, including", "conditions for safe storage, including",
            "equipment and emergency procedures", "and cleaning up", "any incompatibilities",
            "suitable (unsuitable) extinguishing media", "(unsuitable) extinguishing media",
            "specific hazards arising from the chemical", "specific hazards", "from the chemical", 
            "special protective actions for firefighters", "special protective", "for firefighters",
            "handling", "incompatible materials", "safe storage", "contact with",
            ", including any incompatibilities", "including any incompatibilities"
        ]
        sensitive_garbage_regex = []
    elif mode == "HP(K)":
        garbage_heads = ["에 접촉했을 때", "에 들어갔을 때", "들어갔을 때", "접촉했을 때", "했을 때", "흡입했을 때", "먹었을 때", "주의사항", "내용물", "취급요령", "저장방법", "보호구", "조치사항", "제거 방법", "소화제", "유해성", "로부터 생기는", "착용할 보호구", "예방조치", "방법", "경고표지 항목", "그림문자", "화학물질", "의사의 주의사항", "기타 의사의 주의사항", "필요한 정보", "관한 정보", "보호하기 위해 필요한 조치사항", "또는 제거 방법", "시 착용할 보호구 및 예방조치", "시 착용할 보호구", "부터 생기는 특정 유해성", "사의 주의사항", "(부적절한) 소화제", "및", "요령", "때", "항의", "색상", "인화점", "비중", "굴절률", "에 의한 규제", "의한 규제", "- 색", "(및 부적절한) 소화제", "특정 유해성", "보호하기 위해 필요한 조치 사항 및 보호구", "저장 방법"]
        sensitive_garbage_regex = [r"^시\s+", r"^또는\s+", r"^의\s+"]
    else: 
        garbage_heads = ["에 접촉했을 때", "에 들어갔을 때", "들어갔을 때", "접촉했을 때", "했을 때", "흡입했을 때", "먹었을 때", "주의사항", "내용물", "취급요령", "저장방법", "보호구", "조치사항", "제거 방법", "소화제", "유해성", "로부터 생기는", "착용할 보호구", "예방조치", "방법", "경고표지 항목", "그림문자", "화학물질", "의사의 주의사항", "기타 의사의 주의사항", "필요한 정보", "관한 정보", "보호하기 위해 필요한 조치사항", "또는 제거 방법", "시 착용할 보호구 및 예방조치", "시 착용할 보호구", "부터 생기는 특정 유해성", "사의 주의사항", "(부적절한) 소화제", "및", "요령", "때", "항의", "색상", "인화점", "비중", "굴절률", "에 의한 규제", "의한 규제"]
        sensitive_garbage_regex = [r"^시\s+", r"^또는\s+", r"^의\s+"]

    cleaned_lines = []
    for line in target_lines:
        txt = line['text'].strip()
        if mode == "HP(K)": txt = re.sub(r'^\s*-\s*', '', txt).strip()
        
        for _ in range(3):
            changed = False
            for gb in garbage_heads:
                if txt.lower().replace(" ","").startswith(gb.lower().replace(" ","")):
                      p = re.compile(r"^" + re.escape(gb).replace(r"\ ", r"\s*") + r"[\s\.:]*", re.IGNORECASE)
                      m = p.match(txt)
                      if m: txt = txt[m.end():].strip(); changed = True
                      elif txt.lower().startswith(gb.lower()): txt = txt[len(gb):].strip(); changed = True
            
            for pat in sensitive_garbage_regex:
                m = re.search(pat, txt)
                if m: txt = txt[m.end():].strip(); changed = True
            
            txt = re.sub(r"^[:\.\)\s]+", "", txt)
            if not changed: break
        
        if txt:
            if mode == "HP(K)": txt = re.sub(r'^\s*-\s*', '', txt).strip()
            line['text'] = txt
            cleaned_lines.append(line)
    
    if not cleaned_lines: return ""

    final_text = ""
    if len(cleaned_lines) > 0:
        final_text = cleaned_lines[0]['text']
        for i in range(1, len(cleaned_lines)):
            prev = cleaned_lines[i-1]; curr = cleaned_lines[i]
            
            if mode in ["CFF(E)", "HP(E)"]:
                prev_txt = prev['text'].strip()
                curr_txt = curr['text'].strip()
                
                starts_with_bullet = re.match(r"^(\-|•|\*|\d+\.)", curr_txt)
                gap = curr['global_y0'] - prev['global_y1']
                
                if starts_with_bullet or gap >= 3.0:
                    final_text += "\n" + curr_txt
                else:
                    final_text += " " + curr_txt
            else: 
                prev_txt = prev['text'].strip(); curr_txt = curr['text'].strip()
                ends_with_sentence = re.search(r"(\.|시오|음|함|것|임|있음|주의|금지|참조|따르시오|마시오)$", prev_txt)
                starts_with_bullet = re.match(r"^(\-|•|\*|\d+\.|[가-하]\.|\(\d+\))", curr_txt)
                if ends_with_sentence or starts_with_bullet: final_text += "\n" + curr_txt
                else:
                    last_char = prev_txt[-1] if prev_txt else ""
                    first_char = curr_txt[0] if curr_txt else ""
                    is_last_hangul = 0xAC00 <= ord(last_char) <= 0xD7A3
                    is_first_hangul = 0xAC00 <= ord(first_char) <= 0xD7A3
                    gap = curr['global_y0'] - prev['global_y1']
                    if gap < 3.0: 
                        if is_last_hangul and is_first_hangul:
                            need_space = False
                            if last_char in ['을', '를', '이', '가', '은', '는', '의', '와', '과', '에', '로', '서']: need_space = True
                            elif last_char in ['고', '며', '여', '해', '나', '면', '니', '등', '및', '또는', '경우', ',', ')', '속']: need_space = True
                            elif any(curr_txt.startswith(x) for x in ['및', '또는', '(', '참고']): need_space = True
                            if need_space: final_text += " " + curr_txt
                            else: final_text += curr_txt
                        else: final_text += " " + curr_txt
                    else: final_text += "\n" + curr_txt
    return final_text

def parse_sec8_hp_content(text):
    if not text: return "자료없음"
    chunks = text.split("-")
    valid_lines = []
    for chunk in chunks:
        clean_chunk = chunk.strip()
        if not clean_chunk: continue
        if ":" in clean_chunk:
            parts = clean_chunk.split(":", 1)
            name_part = parts[0].strip()
            value_part = parts[1].strip()
            if "해당없음" in value_part: continue 
            name_part = name_part.replace("[", "").replace("]", "").strip()
            value_part = value_part.replace("[", "").replace("]", "").strip()
            final_line = f"{name_part} : {value_part}"
            valid_lines.append(final_line)
        else:
            if "해당없음" not in clean_chunk:
                clean_chunk = clean_chunk.replace("[", "").replace("]", "").strip()
                valid_lines.append(clean_chunk)
    if not valid_lines: return "자료없음"
    return "\n".join(valid_lines)

# --------------------------------------------------------------------------
# [메인 파서]
# --------------------------------------------------------------------------
def parse_pdf_final(doc, mode="CFF(K)"):
    all_lines = get_clustered_lines(doc)
    
    result = {
        "hazard_cls": [], "signal_word": "", "h_codes": [], 
        "p_prev": [], "p_resp": [], "p_stor": [], "p_disp": [],
        "composition_data": [], "sec4_to_7": {}, "sec8": {}, "sec9": {}, "sec14": {}, "sec15": {}
    }
    
    sec9_lines = []
    start_9 = -1; end_9 = -1
    for i, line in enumerate(all_lines):
        if "9. PHYSICAL" in line['text'].upper() or "9. 물리화학" in line['text']: start_9 = i
        if "10. STABILITY" in line['text'].upper() or "10. 안정성" in line['text']: end_9 = i; break
    if start_9 != -1:
        if end_9 == -1: end_9 = len(all_lines)
        sec9_lines = all_lines[start_9:end_9]

    if mode == "HP(E)":
        b19_raw = extract_section_smart(all_lines, "A. GHS Classification", "B. GHS label elements", mode)
        b19_clean = re.sub(r'(?m)^\s*-\s*', '', b19_raw).strip()
        result["hazard_cls"] = [l.strip() for l in b19_clean.split('\n') if l.strip()]

        sig_raw = extract_section_smart(all_lines, "Signal word", "Hazard statement", mode)
        result["signal_word"] = re.sub(r'^(?:[sS]\b)?[\s\-\○•]+', '', sig_raw.strip()).strip()

        h_search_text = extract_section_smart(all_lines, "Hazard statement", "Precautionary statement", mode)
        result["h_codes"] = extract_codes_ordered(h_search_text)

        result["p_prev"] = extract_codes_ordered(extract_section_smart(all_lines, "1) Prevention", "2) Response", mode))
        result["p_resp"] = extract_codes_ordered(extract_section_smart(all_lines, "2) Response", "3) Storage", mode))
        result["p_stor"] = extract_codes_ordered(extract_section_smart(all_lines, "3) Storage", "4) Disposal", mode))
        result["p_disp"] = extract_codes_ordered(extract_section_smart(all_lines, "4) Disposal", "C. Other hazards", mode))

        in_comp = False
        for line in all_lines:
            txt = line['text']
            if "3." in txt and ("성분" in txt or "Composition" in txt or "COMPOSITION" in txt): in_comp=True; continue
            if "4." in txt and ("응급" in txt or "First" in txt or "FIRST" in txt): in_comp=False; break
            
            if in_comp:
                if re.search(r'^\d+\.\d+', txt): continue 
                
                c_val = ""
                cn_val = ""
                
                regex_cas_strict = re.compile(r'\b(\d{2,7}\s*-\s*\d{2}\s*-\s*\d)\b')
                cas_found = regex_cas_strict.findall(txt)
                
                if cas_found:
                    c_val = cas_found[0].replace(" ", "")
                    idx = txt.find(cas_found[0])
                    txt_after_cas = txt[idx + len(cas_found[0]):]
                    
                    m_range = re.search(r'\b(\d+(?:\.\d+)?)\s*(?:-|~)\s*(\d+(?:\.\d+)?)\b', txt_after_cas)
                    if m_range:
                        s, e = m_range.group(1), m_range.group(2)
                        if float(s) <= 100 and float(e) <= 100:
                            if s == "1": s = "0"
                            cn_val = f"{s} ~ {e}"
                    else:
                        m_single = re.search(r'\b(\d+(?:\.\d+)?)\b', txt_after_cas)
                        if m_single:
                            try:
                                v = m_single.group(1)
                                if float(v) <= 100: 
                                    cn_val = v
                            except: pass
                            
                if c_val or cn_val:
                    result["composition_data"].append((c_val, cn_val))

        sec5_lines = []
        for i, line in enumerate(all_lines):
            if "5. FIREFIGHTING" in line['text'].upper(): 
                sec5_lines = all_lines[i:]
                break
        if not sec5_lines: sec5_lines = all_lines

        data = {}
        data["B126"] = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(all_lines, "Eye contact", "Skin contact", mode)).strip()
        data["B127"] = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(all_lines, "Skin contact", "Inhalation contact", mode)).strip()
        data["B128"] = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(all_lines, "Inhalation contact", "Ingestion contact", mode)).strip()
        data["B129"] = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(all_lines, "Ingestion contact", "Delayed and", mode)).strip()

        b132_raw = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(sec5_lines, "Suitable", "Specific hazards", mode)).strip()
        data["B132"] = re.sub(r'(?i)^\s*\(unsuitable\)\s*extinguishing\s*media\s*', '', b132_raw).strip()
        
        b134_raw = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(sec5_lines, "Specific hazards arising", "Special protective", mode)).strip()
        data["B134"] = re.sub(r'(?i)^\s*from\s*the\s*chemical\s*', '', b134_raw).strip()
        
        b136_raw = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(sec5_lines, "Special protective actions", "6. ACCIDENTAL", mode)).strip()
        data["B136"] = re.sub(r'(?i)^\s*for\s*firefighters\s*', '', b136_raw).strip()

        data["B140"] = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(all_lines, "Personal precautions", "Environmental precautions", mode)).strip()
        data["B142"] = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(all_lines, "Environmental precautions", "Methods and materials", mode)).strip()
        data["B144"] = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(all_lines, "Methods and materials for containment", "7. HANDLING", mode)).strip()

        b148_raw = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(all_lines, "Precautions for safe", "Conditions for safe", mode)).strip()
        data["B148"] = re.sub(r'(?i)^\s*handling\s*', '', b148_raw).strip()
        
        b150_raw = re.sub(r'(?m)^\s*-\s*', '', extract_section_smart(all_lines, "Conditions for safe storage", "8. EXPOSURE", mode)).strip()
        data["B150"] = re.sub(r'(?i)^[\s,]*including\s*any\s*incompatibilities\s*', '', b150_raw).strip()
        
        result["sec4_to_7"] = data

        s8_raw = extract_section_smart(all_lines, "ACGIH", "OSHA", mode)
        s8_raw = re.sub(r'(?i)^.*TLV\s*', '', s8_raw).strip()
        s8_clean = re.sub(r'[○•\-\*]+', '', s8_raw).strip()
        
        s8 = {}
        if "Not applicable" in s8_clean or "Not available" in s8_clean or not s8_clean:
            s8["B156"] = "no data available"
            s8["B157"] = ""
            s8["B158"] = ""
        else:
            lines = [l.strip() for l in s8_clean.split('\n') if l.strip()]
            s8["B156"] = lines[0] if len(lines) > 0 else "no data available"
            s8["B157"] = lines[1] if len(lines) > 1 else ""
            s8["B158"] = "\n".join(lines[2:]) if len(lines) > 2 else ""
        result["sec8"] = s8

        s9 = {}
        def find_val_in_sec9(lines, keyword):
            for l in lines:
                if keyword.lower() in l['text'].lower():
                    parts = l['text'].split(keyword, 1)
                    if len(parts) > 1:
                        val = re.sub(r'^[:\s\-\.]+', '', parts[1]).strip()
                        if val: return val
            return ""

        c_val = find_val_in_sec9(sec9_lines, "Color")
        s9["B170"] = c_val.capitalize() if c_val else ""
        
        s9["B176"] = find_val_in_sec9(sec9_lines, "Flash point")
        
        sg_val = find_val_in_sec9(sec9_lines, "Specific gravity")
        g_m = re.search(r'([\d\.]+)', sg_val)
        s9["B183"] = f"{g_m.group(1)} ± 0.010" if g_m else ""
        s9["B189"] = "± 0.005"
        result["sec9"] = s9

        s14 = {}
        un_raw = extract_section_smart(all_lines, "UN No.", "Proper shipping name", mode)
        s14["UN"] = re.sub(r'\D', '', un_raw)
        
        name_raw = extract_section_smart(all_lines, "Proper shipping name", ["C. Hazard Class", "Hazard Class"], mode)
        name_cln = re.sub(r'(?i)proper\s*shipping\s*name', '', name_raw)
        name_cln = re.sub(r'(?i)shipping\s*name', '', name_cln)
        s14["NAME"] = re.sub(r'\([^)]*\)', '', name_cln).replace("-", "").strip()
        
        class_raw = extract_section_smart(all_lines, "C. Hazard Class", ["D. IMDG", "Packing group"], mode)
        class_raw = class_raw.replace("-", "")
        class_match = re.search(r'(\d)', class_raw)
        s14["CLASS"] = class_match.group(1) if class_match else ""

        pg_raw = extract_section_smart(all_lines, "Packing group", "E. Marine pollutant", mode)
        s14["PG"] = pg_raw.replace("-", "").strip()

        env_raw = extract_section_smart(all_lines, "E. Marine pollutant", "F. Special precautions", mode)
        s14["ENV"] = env_raw.replace("-", "").strip()
        
        result["sec14"] = s14

        result["sec15"] = {"DANGER": ""}
        return result

    if mode == "CFF(E)":
        hazard_cls_text = extract_section_smart(all_lines, "2. Hazards identification", "2.2 Labelling", mode)
        
        hazard_cls_text = re.sub(r'(Category\s*\d+[A-Za-z]?)', r'\1\n', hazard_cls_text)
        
        hazard_cls_lines = []
        for line in hazard_cls_text.split('\n'):
            line = line.strip()
            if not line: continue
            
            if "2.1 Classification" in line:
                line = line.replace("2.1 Classification of the substance or", "").replace("mixture", "").strip()
                if not line: continue 
            
            if line.lower() == "mixture": continue
            if line.lower() == "mixture.": continue
            
            hazard_cls_lines.append(line)
        result["hazard_cls"] = hazard_cls_lines

        full_text = "\n".join([l['text'] for l in all_lines])
        m_sig = re.search(r"Signal word\s*[:\-\s]*([A-Za-z]+)", full_text, re.IGNORECASE)
        if m_sig: result["signal_word"] = m_sig.group(1).capitalize()
        
        h_search_text = extract_section_smart(all_lines, "2. Hazards", "3. Composition", mode)
        regex_code = re.compile(r"([HP]\s?\d{3}(?:\s*\+\s*[HP]\s?\d{3})*)")
        all_matches = regex_code.findall(h_search_text) 
        seen = set()
        for code_raw in all_matches:
            code = code_raw.replace(" ", "").upper()
            if code in seen: continue
            seen.add(code)
            if code.startswith("H"): result["h_codes"].append(code)
            elif code.startswith("P"):
                p = code.split("+")[0]
                if p.startswith("P2"): result["p_prev"].append(code)
                elif p.startswith("P3"): result["p_resp"].append(code)
                elif p.startswith("P4"): result["p_stor"].append(code)
                elif p.startswith("P5"): result["p_disp"].append(code)

        comp_text = extract_section_smart(all_lines, "3. Composition", "4. FIRST-AID", mode)
        regex_cas = re.compile(r'\b\d{2,7}-\d{2}-\d\b')
        regex_conc = re.compile(r'\b(\d+(?:\.\d+)?)\s*(?:~|-)\s*(\d+(?:\.\d+)?)\b')
        
        cas_list = regex_cas.findall(comp_text)
        conc_list = []
        
        comp_text_no_cas = regex_cas.sub(" ", comp_text)
        for match in regex_conc.finditer(comp_text_no_cas):
            val1 = float(match.group(1))
            val2 = float(match.group(2))
            if val1 <= 100 and val2 <= 100:
                conc_list.append(f"{match.group(1)} ~ {match.group(2)}")
                
        max_len = max(len(cas_list), len(conc_list))
        for i in range(max_len):
            c_val = cas_list[i] if i < len(cas_list) else ""
            cn_val = conc_list[i] if i < len(conc_list) else ""
            result["composition_data"].append((c_val, cn_val))

        data = {}
        data["B125"] = extract_section_smart(all_lines, "4.1 General advice", "4.2 In case of eye contact", mode)
        data["B126"] = extract_section_smart(all_lines, "4.2 In case of eye contact", "4.3 In case of skin contact", mode)
        data["B127"] = extract_section_smart(all_lines, "4.3 In case of skin contact", "4.4 If inhaled", mode)
        data["B128"] = extract_section_smart(all_lines, "4.4 If inhaled", "4.5 If swallowed", mode)
        data["B129"] = extract_section_smart(all_lines, "4.5 If swallowed", "4.6 Special note for doctors", mode)
        if data["B129"]:
            data["B129"] = data["B129"].replace("Medical personnel, and to ensure that take protection measures is recognized for its substance", "")

        data["B132"] = extract_section_smart(all_lines, "5.1 Extinguishing media", "5.2 Special hazards", mode)
        data["B134"] = extract_section_smart(all_lines, "5.2 Special hazards", "5.3 Advice for firefighters", mode)
        if data["B134"]: data["B134"] = data["B134"].replace("substance or mixture", "")

        data["B136"] = extract_section_smart(all_lines, "5.3 Advice for firefighters", "6. Accidental", mode)
        data["B140"] = extract_section_smart(all_lines, "6.1 Personal precautions", "6.2 Environmental", mode)
        if data["B140"]: data["B140"] = data["B140"].replace("equipment and emergency procedures", "")

        data["B142"] = extract_section_smart(all_lines, "6.2 Environmental", "6.3 Methods", mode)
        data["B144"] = extract_section_smart(all_lines, "6.3 Methods", "7. Handling", mode)
        if data["B144"]: data["B144"] = data["B144"].replace("and cleaning up", "")

        data["B148"] = extract_section_smart(all_lines, "7.1 Precautions", "7.2 Conditions", mode)
        data["B150"] = extract_section_smart(all_lines, "7.2 Conditions", "8. Exposure", mode)
        if data["B150"]: data["B150"] = data["B150"].replace("any incompatibilities", "")

        result["sec4_to_7"] = data

        s8 = {}
        s8["B154"] = extract_section_smart(all_lines, "Internal regulations", "ACGIH regulations", mode)
        s8["B156"] = extract_section_smart(all_lines, "ACGIH regulations", "Biological exposure", mode)
        result["sec8"] = s8

        s9 = {}
        s9["B170"] = extract_section_smart(sec9_lines, "Color", "Odor", mode)
        s9["B176"] = extract_section_smart(sec9_lines, "Flash point", "Evaporation rate", mode)
        
        b183_raw = extract_section_smart(sec9_lines, "Specific gravity", "Partition coefficient", mode)
        s9["B183"] = b183_raw.replace("(20/20℃)", "").replace("(Water=1)", "").strip()
        
        b189_raw = extract_section_smart(sec9_lines, "Refractive index", "10. Stability", mode)
        s9["B189"] = b189_raw.replace("(20℃)", "").strip()
        
        result["sec9"] = s9

        s14 = {}
        un_text = extract_section_smart(all_lines, "14.1 UN number", "14.2 Proper", mode)
        s14["UN"] = re.sub(r'\D', '', un_text)
        
        name_text = extract_section_smart(all_lines, "14.2 Proper", "14.3 Transport", mode)
        name_text = re.sub(r'(?i)proper\s*shipping\s*name', '', name_text)
        name_text = re.sub(r'(?i)shipping\s*name', '', name_text)
        s14["NAME"] = re.sub(r'\([^)]*\)', '', name_text).strip()

        class_raw = extract_section_smart(all_lines, "14.3 Transport hazard class", "14.4 Packing group", mode)
        class_match = re.search(r'(\d)', class_raw)
        s14["CLASS"] = class_match.group(1) if class_match else ""

        pg_raw = extract_section_smart(all_lines, "14.4 Packing group", "14.5 Environmental hazard", mode)
        s14["PG"] = pg_raw

        env_raw = extract_section_smart(all_lines, "14.5 Environmental hazard", "IATA", mode)
        s14["ENV"] = env_raw
        
        result["sec14"] = s14

        return result

    if mode == "CFF(K)":
        for i in range(len(all_lines)):
            if "적정선적명" in all_lines[i]['text']:
                target_line = all_lines[i]
                if i > 0:
                    prev_line = all_lines[i-1]
                    if abs(prev_line['global_y0'] - target_line['global_y0']) < 20:
                        if "적정선적명" not in prev_line['text'] and "유엔번호" not in prev_line['text']:
                            all_lines[i]['text'] = target_line['text'] + " " + prev_line['text']
                            all_lines[i-1]['text'] = ""
    
    limit_y = 999999
    for line in all_lines:
        if "3. 구성성분" in line['text'] or "3. 성분" in line['text']:
            limit_y = line['global_y0']; break
    
    full_text_hp = "\n".join([l['text'] for l in all_lines if l['global_y0'] < limit_y])
    
    signal_found = False
    if mode == "HP(K)":
        try:
            start_sig = full_text_hp.find("신호어")
            end_sig = full_text_hp.find("유해", start_sig)
            if start_sig != -1 and end_sig != -1:
                target_area = full_text_hp[start_sig:end_sig]
                m = re.search(r"[-•]\s*(위험|경고)", target_area)
                if m: result["signal_word"] = m.group(1); signal_found = True
        except: pass
    if not signal_found:
        for line in full_text_hp.split('\n'):
            if "신호어" in line:
                val = line.replace("신호어", "").replace(":", "").strip()
                if val in ["위험", "경고"]: result["signal_word"] = val
            elif line.strip() in ["위험", "경고"] and not result["signal_word"]:
                result["signal_word"] = line.strip()
    
    if mode == "HP(K)":
        lines_hp = full_text_hp.split('\n')
        state = 0
        for l in lines_hp:
            if "가. 유해성" in l: state=1; continue
            if "나. 예방조치" in l: state=0; continue
            if state==1 and l.strip():
                if "공급자" not in l and "회사명" not in l:
                    clean_l = l.replace("-", "").strip()
                    if clean_l: result["hazard_cls"].append(clean_l)
    else: 
        lines_hp = full_text_hp.split('\n')
        state = 0
        for l in lines_hp:
            l_ns = l.replace(" ", "")
            if "2.유해성" in l_ns and "위험성" in l_ns: state = 1; continue 
            if "나.예방조치" in l_ns: state = 0; continue
            if state == 1 and l.strip():
                if "가.유해성" in l_ns and "분류" in l_ns:
                    check_header = re.sub(r'[가-하][\.\s]*유해성[\s\.\·ㆍ\-]*위험성[\s\.\·ㆍ\-]*분류[\s:]*', '', l).strip()
                    if not check_header: continue 
                    l = check_header
                if "공급자" not in l and "회사명" not in l:
                    result["hazard_cls"].append(l.strip())

    regex_code = re.compile(r"([HP]\s?\d{3}(?:\s*\+\s*[HP]\s?\d{3})*)")
    all_matches = regex_code.findall(full_text_hp)
    seen = set()
    if "P321" in full_text_hp and "P321" not in all_matches: all_matches.append("P321")
    for code_raw in all_matches:
        code = code_raw.replace(" ", "").upper()
        if code in seen: continue
        seen.add(code)
        if code.startswith("H"): result["h_codes"].append(code)
        elif code.startswith("P"):
            p = code.split("+")[0]
            if p.startswith("P2"): result["p_prev"].append(code)
            elif p.startswith("P3"): result["p_resp"].append(code)
            elif p.startswith("P4"): result["p_stor"].append(code)
            elif p.startswith("P5"): result["p_disp"].append(code)

    regex_conc = re.compile(r'\b(\d+(?:\.\d+)?)\s*(?:~|-)\s*(\d+(?:\.\d+)?)\b')
    regex_cas_strict = re.compile(r'\b(\d{2,7}\s*-\s*\d{2}\s*-\s*\d)\b')
    regex_cas_ec_kill = re.compile(r'\b\d{2,7}\s*-\s*\d{2,3}\s*-\s*\d\b')
    regex_tilde_range = re.compile(r'(\d+(?:\.\d+)?)\s*~\s*(\d+(?:\.\d+)?)') 
    
    in_comp = False
    for line in all_lines:
        txt = line['text']
        if "3." in txt and ("성분" in txt or "Composition" in txt): in_comp=True; continue
        if "4." in txt and ("응급" in txt or "First" in txt): in_comp=False; break
        if in_comp:
            if re.search(r'^\d+\.\d+', txt): continue 
            
            c_val = ""
            cn_val = ""
            
            if mode == "HP(K)":
                cas_found = regex_cas_strict.findall(txt)
                if cas_found:
                    c_val = cas_found[0].replace(" ", "")
                    txt_no_cas = txt.replace(cas_found[0], " " * len(cas_found[0]))
                    m_range = re.search(r'\b(\d+(?:\.\d+)?)\s*(?:-|~)\s*(\d+(?:\.\d+)?)\b', txt_no_cas)
                    if m_range:
                        s, e = m_range.group(1), m_range.group(2)
                        if s == "1": s = "0"
                        cn_val = f"{s} ~ {e}"
                    else:
                        m_single = re.search(r'\b(\d+(?:\.\d+)?)\b', txt_no_cas)
                        if m_single:
                            try:
                                if float(m_single.group(1)) <= 100: cn_val = m_single.group(1)
                            except: pass
            else:
                cas_found = regex_cas_strict.findall(txt)
                if cas_found:
                    c_val = cas_found[0].replace(" ", "")
                else:
                    cas_found_loose = regex_cas_ec_kill.findall(txt)
                    if cas_found_loose:
                        potential_cas = cas_found_loose[0].replace(" ", "")
                        if re.match(r'\d{2,7}-\d{2}-\d', potential_cas): c_val = potential_cas
                
                txt_clean = regex_cas_ec_kill.sub(" ", txt)
                m_tilde = regex_tilde_range.search(txt_clean)
                if m_tilde:
                    s, e = m_tilde.group(1), m_tilde.group(2)
                    if s == "1": s = "0"
                    cn_val = f"{s} ~ {e}"
            
            if c_val or cn_val:
                result["composition_data"].append((c_val, cn_val))

    data = {}
    if mode == "HP(K)":
        data["B125"] = extract_section_smart(all_lines, "가. 눈에", "나. 피부", mode)
        data["B126"] = extract_section_smart(all_lines, "나. 피부", "다. 흡입", mode)
        data["B127"] = extract_section_smart(all_lines, "다. 흡입", "라. 먹었을", mode)
        data["B128"] = extract_section_smart(all_lines, "라. 먹었을", "마. 기타", mode)
        data["B129"] = extract_section_smart(all_lines, "마. 기타", ["5.", "폭발"], mode)
        data["B132"] = extract_section_smart(all_lines, "가. 적절한", "나. 화학물질", mode)
        
        b133_raw = extract_section_smart(all_lines, "나. 화학물질", "다. 화재진압", mode)
        data["B133"] = re.sub(r'^(특정\s*유해성)\s*', '', b133_raw).strip()
        
        data["B134"] = extract_section_smart(all_lines, "다. 화재진압", ["6.", "누출"], mode)
    else: 
        data["B125"] = extract_section_smart(all_lines, "나. 눈", "다. 피부", mode)
        data["B126"] = extract_section_smart(all_lines, "다. 피부", "라. 흡입", mode)
        data["B127"] = extract_section_smart(all_lines, "라. 흡입", "마. 먹었을", mode)
        data["B128"] = extract_section_smart(all_lines, "마. 먹었을", "바. 기타", mode)
        data["B129"] = extract_section_smart(all_lines, "바. 기타", ["5.", "폭발"], mode)
        data["B132"] = extract_section_smart(all_lines, "가. 적절한", "나. 화학물질", mode)
        
        b133_raw = extract_section_smart(all_lines, "나. 화학물질", "다. 화재진압", mode)
        data["B133"] = re.sub(r'^(특정\s*유해성)\s*', '', b133_raw).strip()
        
        data["B134"] = extract_section_smart(all_lines, "다. 화재진압", ["6.", "누출"], mode)
    
    data["B138"] = extract_section_smart(all_lines, "가. 인체를", "나. 환경을", mode)
    data["B139"] = extract_section_smart(all_lines, "나. 환경을", "다. 정화", mode)
    data["B140"] = extract_section_smart(all_lines, "다. 정화", ["7.", "취급"], mode)
    data["B143"] = extract_section_smart(all_lines, "가. 안전취급", "나. 안전한", mode)
    data["B144"] = extract_section_smart(all_lines, "나. 안전한", ["8.", "노출"], mode)
    result["sec4_to_7"] = data

    sec8_lines = []
    start_8 = -1; end_8 = -1
    for i, line in enumerate(all_lines):
        if "8. 노출방지" in line['text']: start_8 = i
        if "9. 물리화학" in line['text']: end_8 = i; break
    if start_8 != -1:
        if end_8 == -1: end_8 = len(all_lines)
        sec8_lines = all_lines[start_8:end_8]
    
    if mode == "HP(K)":
        b148_raw = extract_section_smart(sec8_lines, "국내노출기준", "ACGIH노출기준", mode)
        b150_raw = extract_section_smart(sec8_lines, "ACGIH노출기준", "생물학적", mode)
        b148_raw = parse_sec8_hp_content(b148_raw)
        b150_raw = parse_sec8_hp_content(b150_raw)
    else:
        b148_raw = extract_section_smart(sec8_lines, "국내규정", "ACGIH", mode)
        b150_raw = extract_section_smart(sec8_lines, "ACGIH", "생물학적", mode)
    result["sec8"] = {"B148": b148_raw, "B150": b150_raw}

    if mode == "HP(K)":
        result["sec9"] = {
            "B163": extract_section_smart(sec9_lines, "- 색", "나. 냄새", mode),
            "B169": extract_section_smart(sec9_lines, "인화점", "아. 증발속도", mode),
            "B176": extract_section_smart(sec9_lines, "비중", "거. n-옥탄올", mode),
            "B182": extract_section_smart(sec9_lines, "굴절률", ["10. 안정성", "10. 화학적"], mode)
        }
    elif mode == "CFF(K)":
        result["sec9"] = {
            "B163": extract_section_smart(sec9_lines, "색상", "나. 냄새", mode),
            "B169": extract_section_smart(sec9_lines, "인화점", "아. 증발속도", mode),
            "B176": extract_section_smart(sec9_lines, "비중", "거. n-옥탄올", mode),
            "B182": extract_section_smart(sec9_lines, "굴절률", ["10. 안정성", "10. 화학적"], mode)
        }

    sec14_lines = []
    start_14 = -1; end_14 = -1
    for i, line in enumerate(all_lines):
        if "14. 운송에" in line['text']: start_14 = i
        if "15. 법적규제" in line['text']: end_14 = i; break
    if start_14 != -1:
        if end_14 == -1: end_14 = len(all_lines)
        sec14_lines = all_lines[start_14:end_14]
    
    if mode == "HP(K)":
        un_no = extract_section_smart(sec14_lines, "유엔번호", "나. 유엔", mode)
        ship_name = extract_section_smart(sec14_lines, "유엔 적정 선적명", ["다. 운송에서의", "다.운송에서의"], mode)
        class_raw = extract_section_smart(sec14_lines, "다. 운송에서의 위험성 등급", ["라. 용기등급", "라.용기등급"], mode)
        pg_raw = extract_section_smart(sec14_lines, "라. 용기등급", ["마. 해양오염물질", "마.해양오염물질"], mode)
        env_raw = extract_section_smart(sec14_lines, "마. 해양오염물질", ["바. 사용자", "바.사용자"], mode)
        
        pg_raw = re.sub(r'\(\s*IMDG\s*CODE\s*/\s*IATA\s*DGR\s*\)', '', pg_raw, flags=re.IGNORECASE)
        pg_raw = pg_raw.replace("-", "").strip()
        env_raw = env_raw.replace("-", "").strip()
    else:
        un_no = extract_section_smart(sec14_lines, "유엔번호", "나. 적정선적명", mode)
        ship_name = extract_section_smart(sec14_lines, "적정선적명", ["다. 운송에서의", "다.운송에서의"], mode)
        class_raw = extract_section_smart(sec14_lines, "다. 운송에서의 위험성 등급", ["라. 용기등급", "라.용기등급"], mode)
        pg_raw = extract_section_smart(sec14_lines, "라. 용기등급", "마. 환경유해성", mode)
        env_raw = extract_section_smart(sec14_lines, "마. 환경유해성", "IATA", mode)
        
    if mode in ["HP(K)", "CFF(K)"]:
        class_match = re.search(r'(\d)', class_raw)
        ship_class = class_match.group(1) if class_match else ""
        result["sec14"] = {
            "UN": un_no, 
            "NAME": ship_name, 
            "CLASS": ship_class,
            "PG": pg_raw,
            "ENV": env_raw
        }

    sec15_lines = []
    start_15 = -1; end_15 = -1
    for i, line in enumerate(all_lines):
        if "15. 법적규제" in line['text']: start_15 = i
        if "16. 그 밖의" in line['text']: end_15 = i; break
    if start_15 != -1:
        if end_15 == -1: end_15 = len(all_lines)
        sec15_lines = all_lines[start_15:end_15]
    
    if mode == "HP(K)":
        danger_act = extract_section_smart(sec15_lines, "라. 위험물안전관리법", ["마. 폐기물", "마.폐기물"], mode)
    else:
        danger_act = extract_section_smart(sec15_lines, "위험물안전관리법", ["마. 폐기물", "마.폐기물"], mode)
    result["sec15"] = {"DANGER": danger_act}

    return result

# --------------------------------------------------------------------------
# [4. 메인 실행 구역]
# --------------------------------------------------------------------------
with st.expander("📂 필수 파일 업로드", expanded=True):
    col1, col2 = st.columns(2)
    with col1:
        master_data_file = st.file_uploader("1. 중앙 데이터 (ingredients...xlsx)", type="xlsx")
        loaded_refs, folder_exists = get_reference_images()
        if folder_exists and loaded_refs:
            st.success(f"✅ 기준 그림 {len(loaded_refs)}개 로드됨")
        elif not folder_exists:
            st.warning("⚠️ 'reference_imgs' 폴더 필요")

    with col2:
        template_file = st.file_uploader("2. 양식 파일 (GHS MSDS 양식)", type="xlsx")

product_name_input = st.text_input("제품명 입력")
option = st.selectbox("적용할 양식", ("CFF(K)", "CFF(E)", "HP(K)", "HP(E)"))

refractive_index_input = ""
if option in ["HP(K)", "HP(E)"]:
    refractive_index_input = st.text_input("굴절률 입력")
    
st.write("") 

kor_excel_file = None
kor_form_version = "신버전 (코드 B25~, 물질 80~122행)"

if option in ["CFF(E)", "HP(E)"]:
    st.markdown("---")
    st.markdown("💡 **(선택) 영문(E) 양식 생성 시, 국문 양식에서 코드 및 물질 정보 가져오기**")
    c3, c4 = st.columns(2)
    with c3:
        kor_excel_file = st.file_uploader("3. 국문 엑셀 파일 (선택)", type="xlsx")
    with c4:
        kor_form_version = st.radio("국문 양식 버전 선택", ["신버전 (코드 B25~, 물질 80~122행)", "구버전 (코드 B25~150, 물질 추출)"])

col_left, col_center, col_right = st.columns([4, 2, 4])

if 'converted_files' not in st.session_state:
    st.session_state['converted_files'] = []
    st.session_state['download_data'] = {}

with col_left:
    st.subheader("3. 원본 파일 업로드")
    uploaded_files = st.file_uploader("원본 데이터(PDF)", type=["pdf"], accept_multiple_files=True)

with col_center:
    st.write("") ; st.write("") ; st.write("")
    
    if st.button("▶ 변환 시작", use_container_width=True):
        if uploaded_files and master_data_file and template_file:
            with st.spinner(f"{option} 모드로 변환 중..."):
                
                new_files = []
                new_download_data = {}
                
                code_map = {} 
                cas_name_map = {} 
                kor_data_map = {} 
                eng_data_map = {} 
                
                try:
                    master_data_file.seek(0)
                    file_bytes = master_data_file.read()
                    xls = pd.ExcelFile(io.BytesIO(file_bytes))
                    
                    target_sheet = None
                    for sheet in xls.sheet_names:
                        if "위험" in sheet and "안전" in sheet: target_sheet = sheet; break
                    
                    if target_sheet:
                        df_code = pd.read_excel(io.BytesIO(file_bytes), sheet_name=target_sheet)
                        if "K" in option:
                            target_col_idx = 1
                        else:
                            target_col_idx = 2
                        
                        for _, row in df_code.iterrows():
                            if pd.notna(row.iloc[0]):
                                code_key = str(row.iloc[0]).replace(" ","").upper().strip()
                                val = row.iloc[target_col_idx]
                                code_map[code_key] = str(val).strip() if pd.notna(val) else ""
                    
                    if "K" in option:
                        sheet_kor = None
                        for sheet in xls.sheet_names:
                            if "국문" in sheet: sheet_kor = sheet; break
                        if sheet_kor:
                            df_kor = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_kor)
                            for _, row in df_kor.iterrows():
                                val_cas = row.iloc[0]
                                val_name = row.iloc[1]
                                if pd.notna(val_cas):
                                    c = str(val_cas).replace(" ", "").strip()
                                    n = str(val_name).strip() if pd.notna(val_name) else ""
                                    cas_name_map[c] = n
                                    if n:
                                        kor_data_map[n] = {
                                            'F': str(row.iloc[5]) if len(row) > 5 else "", 
                                            'G': str(row.iloc[6]) if len(row) > 6 else "", 
                                            'H': str(row.iloc[7]) if len(row) > 7 else "",
                                            'P': str(row.iloc[15]) if len(row) > 15 else "", 
                                            'T': str(row.iloc[19]) if len(row) > 19 else "", 
                                            'U': str(row.iloc[20]) if len(row) > 20 else "", 
                                            'V': str(row.iloc[21]) if len(row) > 21 else ""
                                        }
                    else: # E 모드 (CFF E 등)
                        sheet_eng = None
                        for sheet in xls.sheet_names:
                            if "영문" in sheet: sheet_eng = sheet; break
                        if sheet_eng:
                            df_eng = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_eng)
                            for _, row in df_eng.iterrows():
                                val_cas = row.iloc[0]
                                val_name = row.iloc[1]
                                if pd.notna(val_cas):
                                    c = str(val_cas).replace(" ", "").strip()
                                    n = str(val_name).strip() if pd.notna(val_name) else ""
                                    cas_name_map[c] = n
                                    if n:
                                        eng_data_map[n] = {
                                            'F': str(row.iloc[5]) if len(row) > 5 else "", 
                                            'G': str(row.iloc[6]) if len(row) > 6 else "", 
                                            'H': str(row.iloc[7]) if len(row) > 7 else "",
                                            'P': str(row.iloc[15]) if len(row) > 15 else "", 
                                            'Q': str(row.iloc[16]) if len(row) > 16 else "", 
                                            'T': str(row.iloc[19]) if len(row) > 19 else "", 
                                            'U': str(row.iloc[20]) if len(row) > 20 else "", 
                                            'V': str(row.iloc[21]) if len(row) > 21 else ""
                                        }

                except Exception as e:
                    st.error(f"데이터 로드 오류: {e}")

                kor_override_data = None
                if option in ["CFF(E)", "HP(E)"] and kor_excel_file is not None:
                    kor_excel_file.seek(0)
                    kor_wb = load_workbook(io.BytesIO(kor_excel_file.read()), data_only=True)
                    kor_ws = kor_wb.active
                    kor_override_data = {
                        "h_codes": [], "p_prev": [], "p_resp": [], "p_stor": [], "p_disp": [],
                        "composition_data": []
                    }
                    
                    def ext_codes(ws, s_r, e_r, col=2):
                        res = []
                        regex = re.compile(r"([HP]\d{3}(?:\+[HP]\d{3})*)")
                        for r in range(s_r, e_r + 1):
                            if ws.row_dimensions[r].hidden: 
                                continue
                            val = ws.cell(row=r, column=col).value
                            if val:
                                val_str = str(val).strip().upper()
                                if re.match(r'^[HP]\s?\d{3}', val_str):
                                    matches = regex.findall(val_str.replace(" ", ""))
                                    for m in matches:
                                        if m not in res: res.append(m)
                        return res

                    def ext_comp(ws, s_r, e_r, cas_col=4, conc_col=6):
                        res = []
                        cas_regex = re.compile(r'(\d{2,7}\s*-\s*\d{2}\s*-\s*\d)')
                        for r in range(s_r, e_r + 1):
                            if ws.row_dimensions[r].hidden: 
                                continue
                            cas = ws.cell(row=r, column=cas_col).value
                            conc = ws.cell(row=r, column=conc_col).value
                            if cas and str(cas).strip():
                                c_val = str(cas).strip()
                                cn_val = str(conc).strip() if conc else ""
                                match = cas_regex.search(c_val)
                                if match:
                                    c_val_clean = match.group(1).replace(" ", "")
                                    res.append((c_val_clean, cn_val))
                        return res

                    if "신버전" in kor_form_version:
                        kor_override_data["h_codes"] = ext_codes(kor_ws, 25, 36)
                        kor_override_data["p_prev"] = ext_codes(kor_ws, 38, 49)
                        kor_override_data["p_resp"] = ext_codes(kor_ws, 50, 63)
                        kor_override_data["p_stor"] = ext_codes(kor_ws, 64, 69)
                        kor_override_data["p_disp"] = ext_codes(kor_ws, 70, 72)
                        kor_override_data["composition_data"] = ext_comp(kor_ws, 80, 122)
                    else:
                        all_c = ext_codes(kor_ws, 25, 70)
                        kor_override_data["h_codes"] = [c for c in all_c if c.startswith('H')]
                        kor_override_data["p_prev"] = [c for c in all_c if c.startswith('P2')]
                        kor_override_data["p_resp"] = [c for c in all_c if c.startswith('P3')]
                        kor_override_data["p_stor"] = [c for c in all_c if c.startswith('P4')]
                        kor_override_data["p_disp"] = [c for c in all_c if c.startswith('P5')]
                        kor_override_data["composition_data"] = ext_comp(kor_ws, 25, 150)

                for uploaded_file in uploaded_files:
                    try:
                        doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
                        parsed_data = parse_pdf_final(doc, mode=option)
                        
                        if option in ["CFF(E)", "HP(E)"] and kor_override_data:
                            parsed_data["h_codes"] = kor_override_data["h_codes"]
                            parsed_data["p_prev"] = kor_override_data["p_prev"]
                            parsed_data["p_resp"] = kor_override_data["p_resp"]
                            parsed_data["p_stor"] = kor_override_data["p_stor"]
                            parsed_data["p_disp"] = kor_override_data["p_disp"]
                            parsed_data["composition_data"] = kor_override_data["composition_data"]
                        
                        template_file.seek(0)
                        dest_wb = load_workbook(io.BytesIO(template_file.read()))
                        dest_ws = dest_wb.active

                        dest_wb.external_links = []
                        dest_ws._images = []

                        for row in dest_ws.iter_rows():
                            for cell in row:
                                if isinstance(cell, MergedCell): continue
                                if cell.column == 2 and cell.data_type == 'f': cell.value = ""

                        if option == "HP(E)":
                            dest_ws['A50'].alignment = ALIGN_LEFT
                            dest_ws['A64'].alignment = ALIGN_LEFT
                            dest_ws['A70'].alignment = ALIGN_LEFT
                            
                            safe_write_force(dest_ws, 6, 2, product_name_input, center=True)
                            safe_write_force(dest_ws, 9, 2, product_name_input, center=False)
                            
                            if parsed_data["hazard_cls"]:
                                clean_cls = "\n".join(parsed_data["hazard_cls"])
                                safe_write_force(dest_ws, 19, 2, clean_cls, center=False)
                                dest_ws.row_dimensions[19].height = len(parsed_data["hazard_cls"]) * 14.0
                            
                            if parsed_data["signal_word"]:
                                safe_write_force(dest_ws, 23, 2, parsed_data["signal_word"], center=False)

                            fill_fixed_range(dest_ws, 24, 36, parsed_data["h_codes"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 38, 49, parsed_data["p_prev"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 50, 63, parsed_data["p_resp"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 64, 69, parsed_data["p_stor"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 70, 72, parsed_data["p_disp"], code_map, mode=option)
                            
                            fill_composition_data(dest_ws, parsed_data["composition_data"], cas_name_map, mode=option)
                            
                            active_substances = []
                            for c_data in parsed_data["composition_data"]:
                                cas = c_data[0].replace(" ", "").strip()
                                if cas in cas_name_map:
                                    name = cas_name_map[cas]
                                    if name: active_substances.append(name)
                                    
                            sd = parsed_data["sec4_to_7"]
                            cell_map_e = {
                                "B126": sd.get("B126",""), 
                                "B127": sd.get("B127",""), "B128": sd.get("B128",""),
                                "B129": sd.get("B129",""), "B132": sd.get("B132",""),
                                "B134": sd.get("B134",""), "B136": sd.get("B136",""),
                                "B140": sd.get("B140",""), "B142": sd.get("B142",""),
                                "B144": sd.get("B144",""), "B148": sd.get("B148",""),
                                "B150": sd.get("B150",""),
                                "B170": parsed_data["sec9"].get("B170",""),
                                "B176": parsed_data["sec9"].get("B176",""),
                                "B183": parsed_data["sec9"].get("B183",""),
                                "B189": parsed_data["sec9"].get("B189","")
                            }
                            
                            for addr, val in cell_map_e.items():
                                if not val: continue
                                formatted, h = format_and_calc_height_sec47(val, mode=option)
                                r_idx = int(re.search(r'\d+', addr).group())
                                safe_write_force(dest_ws, r_idx, 2, formatted, center=False)
                                dest_ws.row_dimensions[r_idx].height = h

                            s8 = parsed_data["sec8"]
                            if "B156" in s8:
                                safe_write_force(dest_ws, 156, 2, s8["B156"], center=False)
                                
                                safe_write_force(dest_ws, 157, 2, s8["B157"], center=False)
                                if s8["B157"]: dest_ws.row_dimensions[157].hidden = False
                                else: dest_ws.row_dimensions[157].hidden = True
                                
                                safe_write_force(dest_ws, 158, 2, s8["B158"], center=False)
                                if s8["B158"]: dest_ws.row_dimensions[158].hidden = False
                                else: dest_ws.row_dimensions[158].hidden = True
                            
                            fill_regulatory_section(dest_ws, 202, 240, active_substances, eng_data_map, 'F', mode=option)
                            fill_regulatory_section(dest_ws, 242, 279, active_substances, eng_data_map, 'G', mode=option)
                            fill_regulatory_section(dest_ws, 281, 315, active_substances, eng_data_map, 'H', mode=option)
                            fill_regulatory_section(dest_ws, 324, 358, active_substances, eng_data_map, 'P', mode=option)
                            fill_regulatory_section(dest_ws, 360, 395, active_substances, eng_data_map, 'Q', mode=option)
                            fill_regulatory_section(dest_ws, 401, 437, active_substances, eng_data_map, 'T', mode=option)
                            fill_regulatory_section(dest_ws, 439, 478, active_substances, eng_data_map, 'U', mode=option)
                            fill_regulatory_section(dest_ws, 480, 519, active_substances, eng_data_map, 'V', mode=option)

                            if refractive_index_input:
                                safe_write_force(dest_ws, 182, 2, f"{refractive_index_input.strip()} ± 0.005", center=False)

                            s14 = parsed_data["sec14"]
                            
                            un_val = str(s14.get("UN", "")).strip()
                            if not un_val or un_val.lower() == "not applicable": un_val = "no data available"
                            
                            name_val = str(s14.get("NAME", "")).strip()
                            if not name_val or name_val.lower() == "not applicable": name_val = "no data available"
                            
                            class_val = str(s14.get("CLASS", "")).strip()
                            if not class_val or class_val.lower() == "not applicable": class_val = "Not applicable"
                            
                            pg_val = str(s14.get("PG", "")).strip()
                            if not pg_val or pg_val.lower() == "not applicable": pg_val = "Not applicable"
                            
                            env_val = str(s14.get("ENV", "")).strip()
                            if not env_val or env_val.lower() == "not applicable": env_val = "Not applicable"
                            
                            safe_write_force(dest_ws, 531, 2, un_val, center=False)
                            safe_write_force(dest_ws, 532, 2, name_val, center=False)
                            safe_write_force(dest_ws, 533, 2, class_val, center=False)
                            safe_write_force(dest_ws, 534, 2, pg_val, center=False)
                            safe_write_force(dest_ws, 535, 2, env_val, center=False)

                            today_eng = datetime.now().strftime("%d. %b. %Y").upper()
                            safe_write_force(dest_ws, 544, 1, f"16.2 Date of Issue : {today_eng}", center=False)

                            collected_pil_images = []
                            page = doc[0]
                            image_list = doc.get_page_images(0)
                            
                            for img_info in image_list:
                                xref = img_info[0]
                                try:
                                    base_image = doc.extract_image(xref)
                                    pil_img = PILImage.open(io.BytesIO(base_image["image"]))
                                    
                                    if is_blue_dominant(pil_img): continue
                                    w, h = pil_img.size
                                    if not is_square_shaped(w, h): continue

                                    if loaded_refs:
                                        matched_name = find_best_match_name(pil_img, loaded_refs, mode="HP(K)")
                                        if matched_name:
                                            clean_img = loaded_refs[matched_name]
                                            collected_pil_images.append((extract_number(matched_name), clean_img))
                                except: continue
                            
                            unique_images = {}
                            for key, img in collected_pil_images:
                                if key not in unique_images: unique_images[key] = img
                            
                            final_sorted_imgs = [item[1] for item in sorted(unique_images.items(), key=lambda x: x[0])]

                            if final_sorted_imgs:
                                unit_size = 67; icon_size = 60
                                padding_top = 4; padding_left = (unit_size - icon_size) // 2
                                total_width = unit_size * len(final_sorted_imgs)
                                total_height = unit_size
                                merged_img = PILImage.new('RGBA', (total_width, total_height), (255, 255, 255, 0))
                                for idx, p_img in enumerate(final_sorted_imgs):
                                    p_img_resized = p_img.resize((icon_size, icon_size), PILImage.LANCZOS)
                                    merged_img.paste(p_img_resized, ((idx * unit_size) + padding_left, padding_top))
                                
                                img_byte_arr = io.BytesIO()
                                merged_img.save(img_byte_arr, format='PNG')
                                img_byte_arr.seek(0)
                                dest_ws.add_image(XLImage(img_byte_arr), 'B22') 

                        elif option == "CFF(E)":
                            dest_ws['A50'].alignment = ALIGN_LEFT
                            dest_ws['A64'].alignment = ALIGN_LEFT
                            dest_ws['A70'].alignment = ALIGN_LEFT
                            
                            safe_write_force(dest_ws, 6, 2, product_name_input, center=True)
                            safe_write_force(dest_ws, 9, 2, product_name_input, center=False)
                            
                            if parsed_data["hazard_cls"]:
                                clean_cls = "\n".join(parsed_data["hazard_cls"])
                                safe_write_force(dest_ws, 19, 2, clean_cls, center=False)
                                dest_ws.row_dimensions[19].height = len(parsed_data["hazard_cls"]) * 14.0
                            
                            if parsed_data["signal_word"]:
                                safe_write_force(dest_ws, 23, 2, parsed_data["signal_word"], center=False)

                            fill_fixed_range(dest_ws, 24, 36, parsed_data["h_codes"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 38, 49, parsed_data["p_prev"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 50, 63, parsed_data["p_resp"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 64, 69, parsed_data["p_stor"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 70, 72, parsed_data["p_disp"], code_map, mode=option)
                            
                            fill_composition_data(dest_ws, parsed_data["composition_data"], cas_name_map, mode=option)
                            
                            active_substances = []
                            for c_data in parsed_data["composition_data"]:
                                cas = c_data[0].replace(" ", "").strip()
                                if cas in cas_name_map:
                                    name = cas_name_map[cas]
                                    if name: active_substances.append(name)
                                    
                            sd = parsed_data["sec4_to_7"]
                            cell_map_e = {
                                "B125": sd.get("B125",""), "B126": sd.get("B126",""), 
                                "B127": sd.get("B127",""), "B128": sd.get("B128",""),
                                "B129": sd.get("B129",""), "B132": sd.get("B132",""),
                                "B134": sd.get("B134",""), "B136": sd.get("B136",""),
                                "B140": sd.get("B140",""), "B142": sd.get("B142",""),
                                "B144": sd.get("B144",""), "B148": sd.get("B148",""),
                                "B150": sd.get("B150",""),
                                "B170": parsed_data["sec9"].get("B170","").capitalize(),
                                "B176": parsed_data["sec9"].get("B176",""),
                                "B183": parsed_data["sec9"].get("B183",""),
                                "B189": parsed_data["sec9"].get("B189","")
                            }
                            
                            for addr, val in cell_map_e.items():
                                if not val: continue
                                if addr in ["B183", "B189"] and "±" not in val:
                                    num = re.search(r'([\d\.]+)', val)
                                    if num:
                                        suffix = "0.01" if addr == "B183" else "0.005"
                                        val = f"{num.group(1)} ± {suffix}"
                                
                                formatted, h = format_and_calc_height_sec47(val, mode=option)
                                r_idx = int(re.search(r'\d+', addr).group())
                                safe_write_force(dest_ws, r_idx, 2, formatted, center=False)
                                dest_ws.row_dimensions[r_idx].height = h

                            s8 = parsed_data["sec8"]
                            if s8["B154"]:
                                lines = s8["B154"].split('\n')
                                safe_write_force(dest_ws, 154, 2, lines[0].lower() if "no data" in lines[0].lower() else lines[0], center=False)
                                if len(lines) > 1:
                                    safe_write_force(dest_ws, 155, 2, "\n".join(lines[1:]), center=False)
                                    dest_ws.row_dimensions[155].hidden = False
                            
                            if s8["B156"]:
                                lines = s8["B156"].split('\n')
                                safe_write_force(dest_ws, 156, 2, lines[0].lower() if "no data" in lines[0].lower() else lines[0], center=False)
                                if len(lines) > 1:
                                    safe_write_force(dest_ws, 157, 2, "\n".join(lines[1:]), center=False)
                                    dest_ws.row_dimensions[157].hidden = False

                            fill_regulatory_section(dest_ws, 202, 240, active_substances, eng_data_map, 'F', mode=option)
                            fill_regulatory_section(dest_ws, 242, 279, active_substances, eng_data_map, 'G', mode=option)
                            fill_regulatory_section(dest_ws, 281, 315, active_substances, eng_data_map, 'H', mode=option)
                            fill_regulatory_section(dest_ws, 324, 358, active_substances, eng_data_map, 'P', mode=option)
                            fill_regulatory_section(dest_ws, 360, 395, active_substances, eng_data_map, 'Q', mode=option)
                            fill_regulatory_section(dest_ws, 401, 437, active_substances, eng_data_map, 'T', mode=option)
                            fill_regulatory_section(dest_ws, 439, 478, active_substances, eng_data_map, 'U', mode=option)
                            fill_regulatory_section(dest_ws, 480, 519, active_substances, eng_data_map, 'V', mode=option)

                            s14 = parsed_data["sec14"]
                            
                            un_val = str(s14.get("UN", "")).strip()
                            if not un_val or un_val.lower() == "not applicable": un_val = "no data available"
                            
                            name_val = str(s14.get("NAME", "")).strip()
                            if not name_val or name_val.lower() == "not applicable": name_val = "no data available"
                            
                            class_val = str(s14.get("CLASS", "")).strip()
                            if not class_val or class_val.lower() == "not applicable": class_val = "Not applicable"
                            
                            pg_val = str(s14.get("PG", "")).strip()
                            if not pg_val or pg_val.lower() == "not applicable": pg_val = "Not applicable"
                            
                            env_val = str(s14.get("ENV", "")).strip()
                            if not env_val or env_val.lower() == "not applicable": env_val = "Not applicable"
                            
                            safe_write_force(dest_ws, 531, 2, un_val, center=False)
                            safe_write_force(dest_ws, 532, 2, name_val, center=False)
                            safe_write_force(dest_ws, 533, 2, class_val, center=False)
                            safe_write_force(dest_ws, 534, 2, pg_val, center=False)
                            safe_write_force(dest_ws, 535, 2, env_val, center=False)

                            today_eng = datetime.now().strftime("%d. %b. %Y").upper()
                            safe_write_force(dest_ws, 544, 1, f"16.2 Date of Issue : {today_eng}", center=False)
                            
                            collected_pil_images = []
                            page = doc[0]
                            image_list = doc.get_page_images(0)
                            
                            for img_info in image_list:
                                xref = img_info[0]
                                try:
                                    base_image = doc.extract_image(xref)
                                    pil_img = PILImage.open(io.BytesIO(base_image["image"]))
                                    
                                    if loaded_refs:
                                        matched_name = find_best_match_name(pil_img, loaded_refs, mode=option)
                                        if matched_name:
                                            clean_img = loaded_refs[matched_name]
                                            collected_pil_images.append((extract_number(matched_name), clean_img))
                                except: continue
                            
                            unique_images = {}
                            for key, img in collected_pil_images:
                                if key not in unique_images: unique_images[key] = img
                            
                            final_sorted_imgs = [item[1] for item in sorted(unique_images.items(), key=lambda x: x[0])]

                            if final_sorted_imgs:
                                unit_size = 67; icon_size = 60
                                padding_top = 4; padding_left = (unit_size - icon_size) // 2
                                total_width = unit_size * len(final_sorted_imgs)
                                total_height = unit_size
                                merged_img = PILImage.new('RGBA', (total_width, total_height), (255, 255, 255, 0))
                                for idx, p_img in enumerate(final_sorted_imgs):
                                    p_img_resized = p_img.resize((icon_size, icon_size), PILImage.LANCZOS)
                                    merged_img.paste(p_img_resized, ((idx * unit_size) + padding_left, padding_top))
                                
                                img_byte_arr = io.BytesIO()
                                merged_img.save(img_byte_arr, format='PNG')
                                img_byte_arr.seek(0)
                                dest_ws.add_image(XLImage(img_byte_arr), 'B22') 

                        else: # CFF(K) / HP(K)
                            safe_write_force(dest_ws, 7, 2, product_name_input, center=True)
                            safe_write_force(dest_ws, 10, 2, product_name_input, center=False)
                            
                            if parsed_data["hazard_cls"]:
                                clean_hazard_text = "\n".join([line for line in parsed_data["hazard_cls"] if line.strip()])
                                safe_write_force(dest_ws, 20, 2, clean_hazard_text, center=False)
                                dest_ws['B20'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                                valid_lines_count = len([line for line in parsed_data["hazard_cls"] if line.strip()])
                                if valid_lines_count > 0:
                                    dest_ws.row_dimensions[20].height = valid_lines_count * 14.0

                            signal_final = parsed_data["signal_word"] if parsed_data["signal_word"] else ""
                            safe_write_force(dest_ws, 24, 2, signal_final, center=False) 

                            if option == "HP(K)":
                                safe_write_force(dest_ws, 38, 1, "예방", center=False)
                                safe_write_force(dest_ws, 50, 1, "대응", center=False)
                                safe_write_force(dest_ws, 64, 1, "저장", center=False)
                                safe_write_force(dest_ws, 70, 1, "폐기", center=False)

                            fill_fixed_range(dest_ws, 25, 36, parsed_data["h_codes"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 38, 49, parsed_data["p_prev"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 50, 63, parsed_data["p_resp"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 64, 69, parsed_data["p_stor"], code_map, mode=option)
                            fill_fixed_range(dest_ws, 70, 72, parsed_data["p_disp"], code_map, mode=option)

                            fill_composition_data(dest_ws, parsed_data["composition_data"], cas_name_map, mode=option)
                            
                            active_substances = []
                            for c_data in parsed_data["composition_data"]:
                                cas = c_data[0].replace(" ", "").strip()
                                if cas in cas_name_map:
                                    name = cas_name_map[cas]
                                    if name: active_substances.append(name)

                            sec_data = parsed_data["sec4_to_7"]
                            for cell_addr, raw_text in sec_data.items():
                                formatted_txt, row_h = format_and_calc_height_sec47(raw_text, mode=option)
                                try:
                                    col_str = re.match(r"([A-Z]+)", cell_addr).group(1)
                                    row_num = int(re.search(r"(\d+)", cell_addr).group(1))
                                    col_idx = openpyxl.utils.column_index_from_string(col_str)
                                    safe_write_force(dest_ws, row_num, col_idx, "")
                                    if formatted_txt:
                                        safe_write_force(dest_ws, row_num, col_idx, formatted_txt, center=False)
                                        dest_ws.row_dimensions[row_num].height = row_h
                                        try:
                                            cell_a = dest_ws.cell(row=row_num, column=1)
                                            if cell_a.value: cell_a.value = str(cell_a.value).strip()
                                            cell_a.alignment = ALIGN_TITLE
                                        except: pass
                                except Exception as e: pass

                            s8 = parsed_data["sec8"]
                            val148 = s8["B148"].replace("해당없음", "자료없음")
                            lines148 = [l.strip() for l in val148.split('\n') if l.strip()]
                            safe_write_force(dest_ws, 148, 2, ""); safe_write_force(dest_ws, 149, 2, ""); dest_ws.row_dimensions[149].hidden = True
                            if lines148:
                                safe_write_force(dest_ws, 148, 2, lines148[0], center=False)
                                if len(lines148) > 1:
                                    safe_write_force(dest_ws, 149, 2, "\n".join(lines148[1:]), center=False)
                                    dest_ws.row_dimensions[149].hidden = False
                            
                            val150 = s8["B150"].replace("해당없음", "자료없음")
                            val150 = re.sub(r"^규정[:\s]*", "", val150).strip()
                            safe_write_force(dest_ws, 150, 2, val150, center=False)

                            s9 = parsed_data["sec9"]
                            safe_write_force(dest_ws, 163, 2, s9["B163"], center=False)
                            
                            if option == "HP(K)":
                                flash = s9["B169"]
                                flash_num = re.findall(r'([<>]?\s*\d{2,3})', flash)
                                safe_write_force(dest_ws, 169, 2, f"{flash_num[0]}℃" if flash_num else "", center=False)
                            else:
                                flash = s9["B169"]
                                flash_num = re.findall(r'(\d{2,3})', flash)
                                safe_write_force(dest_ws, 169, 2, f"{flash_num[0]}℃" if flash_num else "", center=False)
                            
                            gravity = s9["B176"].replace("(20℃)", "").replace("(물=1)", "")
                            g_match = re.search(r'([\d\.]+)', gravity)
                            safe_write_force(dest_ws, 176, 2, f"{g_match.group(1)} ± 0.01" if g_match else "", center=False)
                            
                            refract = s9["B182"].replace("(20℃)", "")
                            r_match = re.search(r'([\d\.]+)', refract)
                            
                            if option == "HP(K)" and refractive_index_input:
                                safe_write_force(dest_ws, 182, 2, f"{refractive_index_input.strip()} ± 0.005", center=False)
                            else:
                                safe_write_force(dest_ws, 182, 2, f"{r_match.group(1)} ± 0.005" if r_match else "", center=False)

                            fill_regulatory_section(dest_ws, 195, 226, active_substances, kor_data_map, 'F', mode=option)
                            fill_regulatory_section(dest_ws, 228, 260, active_substances, kor_data_map, 'G', mode=option)
                            fill_regulatory_section(dest_ws, 269, 300, active_substances, kor_data_map, 'H', mode=option)
                            fill_regulatory_section(dest_ws, 316, 348, active_substances, kor_data_map, 'P', mode=option)
                            fill_regulatory_section(dest_ws, 353, 385, active_substances, kor_data_map, 'P', mode=option)
                            fill_regulatory_section(dest_ws, 392, 426, active_substances, kor_data_map, 'T', mode=option)
                            fill_regulatory_section(dest_ws, 428, 460, active_substances, kor_data_map, 'U', mode=option)
                            fill_regulatory_section(dest_ws, 465, 497, active_substances, kor_data_map, 'V', mode=option)

                            for r in range(261, 268): dest_ws.row_dimensions[r].hidden = True
                            for r in range(349, 352): dest_ws.row_dimensions[r].hidden = True
                            dest_ws.row_dimensions[386].hidden = True
                            for r in range(461, 464): dest_ws.row_dimensions[r].hidden = True

                            s14 = parsed_data["sec14"]
                            
                            un_raw = str(s14.get("UN", "")).strip()
                            un_val = re.sub(r"\D", "", un_raw)
                            if not un_val or "해당없음" in un_raw: un_val = "자료없음"
                            
                            name_raw = str(s14.get("NAME", "")).strip()
                            name_val = re.sub(r"\([^)]*\)", "", name_raw).strip()
                            if not name_val or "해당없음" in name_raw: name_val = "자료없음"
                            
                            class_val = str(s14.get("CLASS", "")).strip()
                            if not class_val or "해당없음" in class_val: class_val = "해당없음"
                            
                            pg_val = str(s14.get("PG", "")).strip()
                            if not pg_val or "해당없음" in pg_val: pg_val = "해당없음"
                            
                            env_val = str(s14.get("ENV", "")).strip()
                            if not env_val or "해당없음" in env_val: env_val = "해당없음"

                            safe_write_force(dest_ws, 512, 2, un_val, center=False)
                            safe_write_force(dest_ws, 513, 2, name_val, center=False)
                            safe_write_force(dest_ws, 514, 2, class_val, center=False)
                            safe_write_force(dest_ws, 515, 2, pg_val, center=False)
                            safe_write_force(dest_ws, 516, 2, env_val, center=False)

                            s15 = parsed_data["sec15"]
                            
                            danger_text = s15.get("DANGER", "").strip()
                            clean_danger = danger_text.replace("-", "").replace(" ", "").replace("\n", "")
                            target_cff = "4류 제3석유류(비수용성) 2,000L".replace("-", "").replace(" ", "")
                            target_hp = "위험물에 해당됨 : 제4류 인화성액체, 제3석유류 (비수용성액체) (지정수량 : 2,000리터)".replace("-", "").replace(" ", "")
                            
                            if option == "CFF(K)" and clean_danger == target_cff:
                                safe_write_force(dest_ws, 521, 2, "4류 제3석유류(비수용성) 2,000L", center=False)
                            elif option == "HP(K)" and clean_danger == target_hp:
                                safe_write_force(dest_ws, 521, 2, "4류 제3석유류(비수용성) 2,000L", center=False)
                            else:
                                safe_write_force(dest_ws, 521, 2, "", center=False)
                                dest_ws.cell(row=521, column=2).fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

                            today_str = datetime.now().strftime("%Y.%m.%d")
                            safe_write_force(dest_ws, 542, 2, today_str, center=False)

                        # [HP(K) 이미지 복원: 과거 코드 로직 적용]
                        if option in ["CFF(K)", "HP(K)", "CFF(E)"]:
                            collected_pil_images = []
                            page = doc[0]
                            image_list = doc.get_page_images(0)
                            
                            for img_info in image_list:
                                xref = img_info[0]
                                try:
                                    base_image = doc.extract_image(xref)
                                    pil_img = PILImage.open(io.BytesIO(base_image["image"]))
                                    
                                    if option == "HP(K)":
                                        if is_blue_dominant(pil_img): continue
                                        w, h = pil_img.size
                                        if not is_square_shaped(w, h): continue

                                    if loaded_refs:
                                        matched_name = find_best_match_name(pil_img, loaded_refs, mode=option)
                                        if matched_name:
                                            clean_img = loaded_refs[matched_name]
                                            collected_pil_images.append((extract_number(matched_name), clean_img))
                                except: continue
                            
                            unique_images = {}
                            for key, img in collected_pil_images:
                                if key not in unique_images: unique_images[key] = img
                            
                            final_sorted_imgs = [item[1] for item in sorted(unique_images.items(), key=lambda x: x[0])]

                            if final_sorted_imgs:
                                unit_size = 67; icon_size = 60
                                padding_top = 4; padding_left = (unit_size - icon_size) // 2
                                total_width = unit_size * len(final_sorted_imgs)
                                total_height = unit_size
                                merged_img = PILImage.new('RGBA', (total_width, total_height), (255, 255, 255, 0))
                                for idx, p_img in enumerate(final_sorted_imgs):
                                    p_img_resized = p_img.resize((icon_size, icon_size), PILImage.LANCZOS)
                                    merged_img.paste(p_img_resized, ((idx * unit_size) + padding_left, padding_top))
                                
                                img_byte_arr = io.BytesIO()
                                merged_img.save(img_byte_arr, format='PNG')
                                img_byte_arr.seek(0)
                                dest_ws.add_image(XLImage(img_byte_arr), 'B22') 

                        dest_wb.external_links = []
                        output = io.BytesIO()
                        dest_wb.save(output)
                        output.seek(0)
                        
                        final_name = f"{product_name_input} GHS MSDS({'E' if 'E' in option else 'K'}).xlsx"
                        if final_name in new_download_data:
                            final_name = f"{product_name_input}_{uploaded_file.name.split('.')[0]}.xlsx"
                        new_download_data[final_name] = output.getvalue()
                        new_files.append(final_name)
                        
                    except Exception as e:
                        st.error(f"오류 ({uploaded_file.name}): {e}")

            st.session_state['converted_files'] = new_files
            st.session_state['download_data'] = new_download_data
            
            if 'df_code' in locals(): del df_code
            if 'doc' in locals(): doc.close()
            if 'dest_wb' in locals(): del dest_wb
            if 'output' in locals(): del output
            gc.collect()

            if new_files: st.success("완료!")
    else:
        st.error("모든 파일을 업로드해주세요.")

with col_right:
    st.subheader("결과 다운로드")
    if st.session_state['converted_files']:
        for i, fname in enumerate(st.session_state['converted_files']):
            c1, c2 = st.columns([3, 1])
            with c1: st.text(f"📄 {fname}")
            with c2:
                st.download_button(
                    label="받기", 
                    data=st.session_state['download_data'][fname], 
                    file_name=fname, 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=i
                )
